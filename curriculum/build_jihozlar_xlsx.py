# -*- coding: utf-8 -*-
"""
jihozlar.py dagi ro'yxatdan yuklab olinadigan Excel faylni yasaydi.

Natija: site/jihozlar_5-8-sinf.xlsx  (sayt ildizida — brauzerdan yuklab olinadi)

Ishga tushirish (curriculum/ papkasidan):
    python build_jihozlar_xlsx.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from jihozlar import SET_A, set_b_full, UMUMIY, SET_B_EXTRA

HERE = os.path.dirname(os.path.abspath(__file__))
SITE = os.path.dirname(HERE)          # curriculum/ sayt ildizining ichida
OUT = os.path.join(SITE, "jihozlar_5-8-sinf.xlsx")

GREEN = "17602D"      # brend yashili
GREEN_LIGHT = "EAF1EB"
GREY = "7A8B80"

thin = Side(style="thin", color="D9E4DA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["№", "Kategoriya", "Komponent", "Soni", "Birlik", "Nima uchun / qaysi mavzu"]
WIDTHS = [5, 20, 42, 8, 10, 52]


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor=GREEN)
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 26
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_sheet(wb, title, rows, intro):
    ws = wb.create_sheet(title)

    # Izoh qatori
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    c = ws.cell(row=1, column=1, value=intro)
    c.font = Font(bold=True, size=11, color=GREEN)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34

    style_header(ws, row=2)

    prev_cat = None
    r = 3
    for i, (cat, nom, soni, birlik, mavzu) in enumerate(rows, start=1):
        shade = (cat != prev_cat)
        prev_cat = cat
        vals = [i, cat, nom, soni, birlik, mavzu]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="center" if j in (1, 4, 5) else "left",
                vertical="center", wrap_text=(j == 6))
            if shade:
                cell.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
            if j == 3:
                cell.font = Font(size=10.5, bold=shade)
        r += 1

    # Jami qatori
    ws.cell(row=r, column=3, value="JAMI pozitsiya:").font = Font(bold=True)
    ws.cell(row=r, column=4, value=len(rows)).font = Font(bold=True, color=GREEN)
    ws.cell(row=r, column=5, value="tur").font = Font(bold=True)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:F{r-1}"
    return ws


def build():
    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(
        wb, "SET A — 5-6-sinf", SET_A,
        "SET A — 5-6-sinf uchun to'plam (elektronika asoslari + Arduino, blokli dasturlash). "
        "Miqdorlar BITTA to'plam uchun. 1 to'plam = 2 o'quvchi.")

    write_sheet(
        wb, "SET B — 7-8-sinf", set_b_full(),
        "SET B — 7-8-sinf uchun to'plam (ilg'or Arduino + ESP32 + AI). "
        "SET A ning hammasini o'z ichiga oladi va ustiga qo'shimchalar. "
        "Miqdorlar BITTA to'plam uchun. 1 to'plam = 2 o'quvchi.")

    write_sheet(
        wb, "SET B — faqat qo'shimcha", SET_B_EXTRA,
        "Agar SET A allaqachon sotib olingan bo'lsa — 7-8-sinfga o'tish uchun "
        "FAQAT shu ro'yxatni qo'shib olish yetarli.")

    write_sheet(
        wb, "Sinf uchun umumiy", UMUMIY,
        "Bu jihozlar har to'plamga emas, BUTUN SINFGA olinadi "
        "(kavsharlash, asboblar, xavfsizlik, zaxira).")

    # --- Xulosa varag'i ---
    ws = wb.create_sheet("Xulosa", 0)
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 60

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="5-8-SINF ROBOTOTEXNIKA KURSI — JIHOZ RO'YXATI")
    t.font = Font(bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=GREEN)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    rows = [
        ("", "", ""),
        ("To'plam", "Pozitsiya", "Kim uchun / nima beradi"),
        ("SET A", len(SET_A),
         "5-6-sinf: elektronika asoslari, o'lchov, blokli dasturlash, oddiy sensorlar"),
        ("SET B", len(set_b_full()),
         "7-8-sinf: SET A + ESP32, kamera/mikrofon, AI sensorlari, IoT"),
        ("SET B qo'shimchasi", len(SET_B_EXTRA),
         "SET A bor bo'lsa — 7-8-sinfga o'tish uchun shuni qo'shish kifoya"),
        ("Sinf uchun umumiy", len(UMUMIY),
         "Kavsharlash stansiyasi, asboblar, xavfsizlik, zaxira platalar"),
        ("", "", ""),
        ("MUHIM IZOHLAR", "", ""),
        ("Multimetr", "majburiy",
         "Arduino rasmiy dasturida Om qonuni darsi to'liq o'lchovga qurilgan. "
         "Multimetrsiz bu dars o'tilmaydi."),
        ("XIAO ESP32S3 Sense", "AI uchun asosiy",
         "Kamera va mikrofon plataning o'zida. AI (Edge Impulse/TinyML) uchun "
         "tasvir va ovoz ma'lumoti shundan olinadi."),
        ("MPU6050", "AI uchun",
         "Imo-ishorani tanish modeli uchun harakat ma'lumoti beradi."),
        ("Nechta to'plam kerak", "o'quvchi/2",
         "12 o'quvchilik sinf uchun 6 to'plam. Yakka ishlatilsa — 12 to'plam."),
    ]
    r = 2
    for a, b, c in rows:
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        cc = ws.cell(row=r, column=3, value=c)
        if a in ("To'plam", "MUHIM IZOHLAR"):
            for cell in (ca, cb, cc):
                cell.font = Font(bold=True, color=GREEN)
                cell.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
        cc.alignment = Alignment(wrap_text=True, vertical="center")
        ca.alignment = Alignment(vertical="center")
        cb.alignment = Alignment(horizontal="center", vertical="center")
        if c:
            ws.row_dimensions[r].height = 30
        r += 1

    wb.save(OUT)
    print("Yozildi:", OUT)
    print("SET A:", len(SET_A), "pozitsiya")
    print("SET B:", len(set_b_full()), "pozitsiya (qo'shimcha:", len(SET_B_EXTRA), ")")
    print("Sinf uchun umumiy:", len(UMUMIY), "pozitsiya")


if __name__ == "__main__":
    build()
