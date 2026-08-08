# -*- coding: utf-8 -*-
import math
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from curriculum_data import ALL_GRADES_Y1, GRADE4_Y2
from year1_model_pools import YEAR1_POOLS
from models_catalog import CATALOG
from steam_notes import get_steam_note
from amaliy_tags import AMALIY
from dasturlash_track import DASTURLASH_TRACK

FONT = "Arial"
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
MOD_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_title(cell, size=13):
    cell.font = Font(name=FONT, bold=True, size=size, color="FFFFFF")
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(vertical="center")

def style_note(cell):
    cell.font = Font(name=FONT, italic=True, size=10, color="404040")

def style_section(cell):
    cell.font = Font(name=FONT, bold=True, size=11)

def style_header(cell):
    cell.font = Font(name=FONT, bold=True, size=10)
    cell.fill = HEADER_FILL
    cell.border = BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)

def style_cell(cell, bold=False, wrap=True, fill=None):
    cell.font = Font(name=FONT, bold=bold, size=10)
    cell.border = BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = fill

def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def extract_amaliy(topic, modcode, grade_key):
    """1-yil / 2-yil qurish darslari uchun amaliy natijani dars mavzusidan chiqarib olish."""
    if " — yig'ish va sinov (1 darslik loyiha)" in topic:
        return topic.split(" — yig'ish va sinov")[0]
    if topic.startswith("Chorak kirish"):
        return "Namoyish / tanishuv (keyingi darsdan qurish boshlanadi)"
    if modcode.startswith("N") and '"' in topic:
        race_name = topic.split('"')[1]
        return f"Musobaqa: {race_name} (robot bilan ishtirok etish)"
    if modcode.startswith("N"):
        return "Amaliy nazorat: 2 model tanlab, yig'ib va tushuntirib berish"
    if modcode.startswith("L"):
        return "Erkin loyiha: shu chorak modellaridan birini o'zgartirib qurish"
    return "—"

def write_grade_sheet(ws, title, note_line, grade_data, grade_key=None, extra_note=None):
    ws["A1"] = title
    ws.merge_cells("A1:F1")
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    ws["A2"] = note_line
    ws.merge_cells("A2:F2")
    style_note(ws["A2"])

    row = 3
    if extra_note:
        ws.cell(row=row, column=1, value=extra_note)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        style_note(ws.cell(row=row, column=1))
        row += 1

    row += 1  # blank
    ws.cell(row=row, column=1, value="Modullar (QURISH — dasturlashsiz)")
    style_section(ws.cell(row=row, column=1))
    row += 1

    hdr_row = row
    headers = ["Modul", "Darslar soni", "Chorak", "Asosiy jihoz", "Bu chorakda quriladigan modellar (18 ta)"]
    for c, h in enumerate(headers, 1):
        style_header(ws.cell(row=hdr_row, column=c, value=h))
    row += 1

    module_count_cells = []
    lesson_rows = []  # (num, hafta, modul_code, dars_mavzusi, amaliy_ish, qollanma)
    lesson_num = 0

    pool = YEAR1_POOLS.get(grade_key) if grade_key else None
    is_spike = grade_key is None
    amaliy_data = AMALIY.get("4-sinf-spike") if is_spike else None

    for qi, q in enumerate(grade_data["quarters"], 1):
        models_str = ", ".join(pool[qi - 1]) if pool else "-"
        # content module row
        ws.cell(row=row, column=1, value=q["nomi"])
        ws.cell(row=row, column=2, value=19)
        ws.cell(row=row, column=3, value=f"{qi}-chorak")
        ws.cell(row=row, column=4, value=q["jihoz"])
        ws.cell(row=row, column=5, value=models_str)
        for c in range(1, 6):
            style_cell(ws.cell(row=row, column=c), fill=MOD_FILL if c == 1 else None)
        module_count_cells.append(row)
        row += 1
        # nazorat row
        ws.cell(row=row, column=1, value=f"N{qi}. Nazorat ishi")
        ws.cell(row=row, column=2, value=1)
        ws.cell(row=row, column=3, value=f"{qi}-chorak")
        ws.cell(row=row, column=4, value="-")
        ws.cell(row=row, column=5, value="-")
        for c in range(1, 6):
            style_cell(ws.cell(row=row, column=c))
        module_count_cells.append(row)
        row += 1
        # loyiha row
        ws.cell(row=row, column=1, value=f"L{qi}. Loyiha")
        ws.cell(row=row, column=2, value=1)
        ws.cell(row=row, column=3, value=f"{qi}-chorak")
        ws.cell(row=row, column=4, value=q["jihoz"])
        ws.cell(row=row, column=5, value=models_str)
        for c in range(1, 6):
            style_cell(ws.cell(row=row, column=c))
        module_count_cells.append(row)
        row += 1

        # build lesson rows for dars-mavzu table
        qam = amaliy_data[qi - 1] if amaliy_data else None
        for li, item in enumerate(q["lessons"]):
            lesson_num += 1
            hafta = math.ceil(lesson_num / 2)
            if isinstance(item, dict):
                # YANGI format: STEAM tema (Dars mavzusi) + model (Amaliy ish) alohida
                dars_mavzusi = item["mavzu"]
                amaliy_ish = item["model"] if item["model"] else "—"
                qollanma = item["guide"] if item["guide"] else (
                    "Yangi mavzu bilan tanishtiring, keyingi darsdan qurish boshlanadi.")
            else:
                # ESKI format (SPIKE): oddiy matn
                topic = item
                dars_mavzusi = topic
                amaliy_ish = qam["lessons"][li] if qam else extract_amaliy(topic, f"M{qi}", grade_key)
                qollanma = get_steam_note(topic)
            lesson_rows.append((lesson_num, hafta, f"M{qi}", dars_mavzusi, amaliy_ish, qollanma))

        lesson_num += 1
        hafta = math.ceil(lesson_num / 2)
        if qam:
            amaliy_n = qam["nazorat"]
        else:
            amaliy_n = extract_amaliy(q["nazorat"], f"N{qi}", grade_key)
        qollanma_n = ("Musobaqa formatida o'tkaziladi: xronometr/soat bilan vaqtni o'lchang, "
                      "natijani jadvalga yozing, g'oliblarni e'lon qiling. Baholash mezoni "
                      "\"Dars mavzusi\" ustunida ko'rsatilgan.")
        lesson_rows.append((lesson_num, hafta, f"N{qi}", q["nazorat"], amaliy_n, qollanma_n))

        lesson_num += 1
        hafta = math.ceil(lesson_num / 2)
        if qam:
            amaliy_l = qam["loyiha"]
        else:
            amaliy_l = extract_amaliy(q["loyiha"], f"L{qi}", grade_key)
        qollanma_l = "Erkin loyiha: o'quvchi tanlagan modelni tushuntirib beradi."
        lesson_rows.append((lesson_num, hafta, f"L{qi}", q["loyiha"], amaliy_l, qollanma_l))

    # Jami rejali darslar (SUM formula)
    first_mc = module_count_cells[0]
    last_mc = module_count_cells[-1]
    ws.cell(row=row, column=1, value="Jami QURISH darslari")
    style_cell(ws.cell(row=row, column=1), bold=True)
    ws.cell(row=row, column=2, value=f"=SUM(B{first_mc}:B{last_mc})")
    style_cell(ws.cell(row=row, column=2), bold=True)
    style_cell(ws.cell(row=row, column=3))
    style_cell(ws.cell(row=row, column=4))
    style_cell(ws.cell(row=row, column=5))
    qurish_total_row = row
    row += 2

    ws.cell(row=row, column=1, value="Dars-mavzu rejasi — QURISH")
    style_section(ws.cell(row=row, column=1))
    row += 1

    hdr2 = row
    headers2 = ["№", "Hafta", "Modul", "Dars mavzusi (nazariy/STEAM tema)",
                "Amaliy ish (nima yasaladi)", "O'qituvchi uchun qo'llanma (nimalarni tushuntirish kerak)"]
    for c, h in enumerate(headers2, 1):
        style_header(ws.cell(row=hdr2, column=c, value=h))
    row += 1

    for num, hafta, modcode, dars_mavzusi, amaliy_ish, qollanma in lesson_rows:
        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=hafta)
        ws.cell(row=row, column=3, value=modcode)
        ws.cell(row=row, column=4, value=dars_mavzusi)
        ws.cell(row=row, column=5, value=amaliy_ish)
        ws.cell(row=row, column=6, value=qollanma)
        for c in range(1, 7):
            fill = None
            if modcode.startswith("N"):
                fill = PatternFill("solid", fgColor="FFF2CC")
            elif modcode.startswith("L"):
                fill = PatternFill("solid", fgColor="E2EFDA")
            style_cell(ws.cell(row=row, column=c), fill=fill)
        row += 1

    autosize(ws, [42, 9, 12, 30, 30, 55])
    ws.freeze_panes = "A2"
    return row


def write_dasturlash_section(ws, start_row, grade_key):
    """2-yil, 3-4-chorak: YANGI dasturlash kursi (haftaning 3-soati) — qurish jadvalidan keyin qo'shiladi."""
    row = start_row + 1
    ws.cell(row=row, column=1,
            value="QO'SHIMCHA KURS — DASTURLASH (2-yil, 3-4-chorak, HAFTANING 3-SOATI, Scratch-uslub)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    style_title(ws.cell(row=row, column=1), size=12)
    ws.row_dimensions[row].height = 20
    row += 1

    ws.cell(row=row, column=1,
            value="Qurish darslari (yuqorida, 2 soat/hafta) o'zgarishsiz davom etadi. Bu — QO'SHIMCHA "
                  "3-soat: 1-yilda yasalgan modellarga dastur yozish orqali Scratch-uslub blok-dasturlash "
                  "o'rganiladi. Jihoz xarid qilinmaydi — faqat ilova (planshet/telefon).")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
    style_note(ws.cell(row=row, column=1))
    ws.row_dimensions[row].height = 30
    row += 2

    track = DASTURLASH_TRACK[grade_key]
    hdr = row
    headers = ["Bosqich", "Darslar soni", "Chorak", "Jihoz"]
    for c, h in enumerate(headers, 1):
        style_header(ws.cell(row=hdr, column=c, value=h))
    row += 1

    mod_rows = []
    lesson_rows = []
    dars_num = 0
    for qname in ["Q3", "Q4"]:
        qdata = track[qname]
        chorak_no = "3-chorak" if qname == "Q3" else "4-chorak"
        ws.cell(row=row, column=1, value=f"D-{qname}. Dasturlash mavzulari")
        ws.cell(row=row, column=2, value=len(qdata["lessons"]))
        ws.cell(row=row, column=3, value=chorak_no)
        ws.cell(row=row, column=4, value="Scratch-uslub ilova")
        for c in range(1, 5):
            style_cell(ws.cell(row=row, column=c), fill=MOD_FILL if c == 1 else None)
        mod_rows.append(row)
        row += 1
        ws.cell(row=row, column=1, value=f"D-{qname}. Nazorat ishi")
        ws.cell(row=row, column=2, value=1)
        ws.cell(row=row, column=3, value=chorak_no)
        ws.cell(row=row, column=4, value="-")
        for c in range(1, 5):
            style_cell(ws.cell(row=row, column=c))
        mod_rows.append(row)
        row += 1
        ws.cell(row=row, column=1, value=f"D-{qname}. Loyiha")
        ws.cell(row=row, column=2, value=1)
        ws.cell(row=row, column=3, value=chorak_no)
        ws.cell(row=row, column=4, value="Scratch-uslub ilova")
        for c in range(1, 5):
            style_cell(ws.cell(row=row, column=c))
        mod_rows.append(row)
        row += 1

        for topic in qdata["lessons"]:
            dars_num += 1
            lesson_rows.append((dars_num, f"D{qname[-1]}", topic, topic))
        dars_num += 1
        lesson_rows.append((dars_num, f"DN{qname[-1]}", qdata["nazorat"], qdata["nazorat"]))
        dars_num += 1
        lesson_rows.append((dars_num, f"DL{qname[-1]}", qdata["loyiha"], qdata["loyiha"]))

    first_mc, last_mc = mod_rows[0], mod_rows[-1]
    ws.cell(row=row, column=1, value="Jami DASTURLASH darslari")
    style_cell(ws.cell(row=row, column=1), bold=True)
    ws.cell(row=row, column=2, value=f"=SUM(B{first_mc}:B{last_mc})")
    style_cell(ws.cell(row=row, column=2), bold=True)
    style_cell(ws.cell(row=row, column=3))
    style_cell(ws.cell(row=row, column=4))
    dasturlash_total_row = row
    row += 2

    ws.cell(row=row, column=1, value="Dars-mavzu rejasi — DASTURLASH")
    style_section(ws.cell(row=row, column=1))
    row += 1
    hdr2 = row
    for c, h in enumerate(["№", "Bosqich", "Dars mavzusi", "Amaliy natija (nima dasturlanadi)"], 1):
        style_header(ws.cell(row=hdr2, column=c, value=h))
    row += 1
    for num, code, topic, amaliy in lesson_rows:
        is_naz = code.startswith("DN")
        is_loy = code.startswith("DL")
        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=code)
        ws.cell(row=row, column=3, value=topic)
        ws.cell(row=row, column=4, value=amaliy)
        fill = PatternFill("solid", fgColor="FFF2CC") if is_naz else (
            PatternFill("solid", fgColor="E2EFDA") if is_loy else None)
        for c in range(1, 5):
            style_cell(ws.cell(row=row, column=c), fill=fill)
        row += 1

    return row, dasturlash_total_row


def build():
    wb = Workbook()
    wb.remove(wb.active)

    # ---------------- Umumiy reja ----------------
    ov = wb.create_sheet("Umumiy reja")
    ov["A1"] = "ROBOTOTEXNIKA O'QUV DASTURI — 0-4-SINF (MAKERZOID, 1-YIL VA 2-YIL)"
    ov.merge_cells("A1:E1")
    style_title(ov["A1"], size=13)
    ov.row_dimensions[1].height = 24

    ov["A2"] = ("YANGI TUZILMA: 1-yil = 100% QURISH (dasturlash yo'q), haftasiga 2 dars, 38 hafta = "
                "84 dars/yil, har chorakda 18 ta TURLI robot (har biri 1 DARSLIK: yig'ish+sinov bir darsda) = yiliga 72 ta "
                "turli robot/sinf. 2-yil, 1-2-chorak — xuddi shunday qurish davom etadi. 2-yil, "
                "3-4-chorak — haftasiga 3-SOATGA o'tiladi: 2 soat qurish (o'zgarishsiz) + YANGI 1 soat "
                "dasturlash (Scratch-uslub, Makerzoid ilovasi, 1-yilda yasalgan modellarga dastur yoziladi). "
                "4-sinf 2-yilda SPIKE Prime'ga o'tadi (qurish+dasturlash tabiiy birga).")
    ov.merge_cells("A2:E2")
    ov["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    style_note(ov["A2"])
    ov.row_dimensions[2].height = 60

    row = 4
    ov.cell(row=row, column=1, value="Bitta o'quvchi nechta TURLI robot yasaydi?")
    style_section(ov.cell(row=row, column=1))
    ov.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    ov.cell(row=row, column=1,
            value="1 sinf, 1-yil ichida: 72 ta turli robot (4 chorak x 18 ta, har biri BITTA darsda "
                  "to'liq tugaydi — jihozlar keyingi darsga saqlanmaydi). Agar bola 0-sinfdan "
                  "4-sinfgacha (5 yil) davom etsa: 5 x 72 = 360 ta turli robot — bundan tashqari har "
                  "chorakdagi ERKIN LOYIHA (4/yil) alohida hisoblanadi. 2-yilning 3-4-choragida "
                  "qo'shimcha 19 ta dasturlash mashqi (mavjud modellarga dastur yozish) qo'shiladi.")
    ov.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ov.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
    ov.cell(row=row, column=1).font = Font(name=FONT, size=10, bold=True)
    ov.row_dimensions[row].height = 45
    row += 2
    ov.cell(row=row, column=1, value="Sinf")
    ov.cell(row=row, column=2, value="1-YIL: platforma")
    ov.cell(row=row, column=3, value="1-yil mazmuni (qisqacha)")
    ov.cell(row=row, column=4, value="2-YIL: platforma")
    ov.cell(row=row, column=5, value="2-yil mazmuni (qisqacha)")
    for c in range(1, 6):
        style_header(ov.cell(row=row, column=c))
    row += 1

    overview_rows = [
        ("0-sinf", "Makerzoid — 100% qurish",
         "4 chorak x 18 turli robot (har biri 1 darsda) = 72 robot/yil, dasturlashsiz",
         "Qurish (1-yil bilan AYNAN bir xil)",
         "Yangi qabul: xuddi shu yo'l. Dasturlash YO'Q (hali erta)."),
        ("1-sinf", "Makerzoid — 100% qurish",
         "4 chorak x 18 turli robot = 72 robot/yil, dasturlashsiz",
         "Qurish (1-yil bilan AYNAN bir xil)",
         "Yangi qabul: xuddi shu yo'l. Dasturlash YO'Q (hali erta)."),
        ("2-sinf", "Makerzoid — 100% qurish",
         "4 chorak x 18 turli robot = 72 robot/yil, dasturlashsiz",
         "Qurish (1-yil bilan bir xil) + 3-4-chorak: +19 dars dasturlash",
         "Dasturlash AYNAN shu sinfdan boshlanadi (Scratch-uslub)"),
        ("3-sinf", "Makerzoid — 100% qurish",
         "4 chorak x 18 turli robot = 72 robot/yil, dasturlashsiz",
         "Qurish (1-yil bilan bir xil) + 3-4-chorak: +19 dars dasturlash",
         "Xuddi shu qurish dasturi + dasturlash davom etadi"),
        ("4-sinf", "Makerzoid — 100% qurish",
         "4 chorak x 18 turli robot = 72 robot/yil, dasturlashsiz, SPIKE'ga tayyorgarlik",
         "SPIKE Prime (YANGI platforma)",
         "1-chorak 100% qurish (LEGO rasmiy modellari), keyin missiyalar (Invention Squad uslubi)"),
    ]
    for r in overview_rows:
        for c, val in enumerate(r, 1):
            style_cell(ov.cell(row=row, column=c, value=val))
        row += 1

    row += 1
    ov.cell(row=row, column=1, value="Jihoz-dastur muvofiqligi va boshqa muhim eslatmalar")
    style_section(ov.cell(row=row, column=1))
    ov.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    notes = [
        "- ASOSIY TAMOYIL (metodist talabi bilan): 1-yil dasturlash TUTMAYDI — 100% qurish. Har chorakda "
        "18 ta turli, TO'LIQ tugallanadigan robot yasaladi — har biri BITTA darsda (yig'ish+sinov shu darsning o'zida, jihozlar keyingi darsga saqlanmaydi). "
        "Bu bolalar qiziqishini doim yangi natija bilan ushlab turadi (uzoq, sekin loyihalar o'rniga "
        "tez-tez tugallanadigan kichik g'alabalar).",
        "- Dasturlash FAQAT 2-sinfdan boshlab, 3-4-choraqda, YANGI qo'shilgan 3-soatda boshlanadi "
        "(Scratch-uslub, Makerzoid ilovasi). 0-1-sinf — 2-yilda ham dasturlashsiz, 1-yil bilan AYNAN "
        "bir xil. 2-3-sinf, 1-2-chorak — hali sof qurish. Bu haftalik soat sonini 2-3-sinf 3-4-choragida "
        "2 dan 3 ga oshiradi (o'quv rejasida hisobga olinishi kerak).",
        "- USTUNLAR AJRATILGAN (metodist talabi bilan): \"Dars mavzusi\" — NAZARIY/ILMIY tema (masalan "
        "\"Ishqalanish kuchi\", \"Richag qonuni\") — bu haqiqiy pedagogik mavzu, darslikka yoziladigan "
        "nom. \"Amaliy ish\" — qaysi model yasalishi (masalan \"Little Lantern 1\") — bu shunchaki loyiha "
        "nomi, katalogdan. \"O'qituvchi uchun qo'llanma\" — darsda NIMALARNI tushuntirish kerakligi "
        "bo'yicha yo'riqnoma (masalan: chang'i yasalganda — ishqalanish kuchi nima, qanday ishlashi, "
        "silliq/notekis sirt farqi tushuntiriladi, keyin qurishga o'tiladi).",
        "- 3-sinf 3-4-chorak va butun 4-sinf (1-yil): YO'RIQNOMA KAMAYADI — model nomi hali ko'rsatiladi, "
        "lekin o'quvchi konstruksiyani ko'proq o'zi ixtiro qilishga undaladi (hayotiy misollar orqali).",
        "- Nazorat ishlari BARCHA sinflarda (0-4) amaliy shaklda o'tkaziladi — yozma test qo'llanilmaydi.",
        "- Makerzoid Robot Master Standard: 1 host controller + 1 motor + 1 sensor + 370+ blok. "
        "3-4-sinfning ba'zi modellari 2 motor+2 sensor talab qiladi — maktabda bu ALLAQACHON mavjud "
        "(2 ta Standard to'plam birlashtirilgan holda), qo'shimcha xarid shart emas.",
        "- Batareyalar (2xAA/to'plam) standart tarkibga kirmaydi — qayta zaryadlanadigan batareya tavsiya etiladi.",
        "- Dasturlash ilovasi telefon/planshetda ishlaydi — to'liq kompyuter xonasi shart emas.",
        "- 2-yil 0-3-sinf QURISH qismi 1-yildagi mos sinf dasturi bilan AYNAN bir xil (grade-anchored "
        "yondashuv); DASTURLASH qismi esa 2-yilga xos YANGI qo'shimcha kurs.",
        "- O'qituvchi bo'linishi tavsiyasi: boshlang'ich (0-2, qurish asoslari) va yuqori (3-4, "
        "kengaytirilgan/musobaqa) uchun alohida o'qituvchi; 2-yil 3-4-choraqda dasturlash uchun "
        "alohida (yoki qo'shimcha tayyorgarlikdan o'tgan) o'qituvchi kerak bo'lishi mumkin.",
        "- Model tanlovi \"Instruksiya katalogi\" varag'idagi 281 ta Robot Master (Premium) modeliga "
        "asoslangan — har bir sinf varag'ida chorak bo'yicha quriladigan modellar ko'rsatilgan.",
    ]
    for n in notes:
        ov.cell(row=row, column=1, value=n)
        ov.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ov.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
        ov.cell(row=row, column=1).font = Font(name=FONT, size=10)
        ov.row_dimensions[row].height = 28
        row += 1

    autosize(ov, [14, 26, 42, 26, 42])

    # ---------------- 1-yil sheets ----------------
    for grade_name, gdata in ALL_GRADES_Y1.items():
        ws = wb.create_sheet(f"1-yil {grade_name}")
        title = f"1-YIL — {grade_name.upper()} — {gdata['platforma']}"
        note = ("Haftasiga 2 dars (soat), 42 o'quv haftasi. 84 dars: 4 chorak x 21 dars "
                "(18 model x 1 dars + 1 kirish = 19 mavzu darsi + 1 nazorat + 1 loyiha). "
                "100% QURISH — dasturlash YO'Q. Nazorat — amaliy.")
        write_grade_sheet(ws, title, note, gdata, grade_key=grade_name)

    # ---------------- 2-yil 0-1-sinf: FAQAT qurish (1-yil bilan AYNAN bir xil, dasturlashsiz) ----------------
    for grade_name in ["0-sinf", "1-sinf"]:
        gdata = ALL_GRADES_Y1[grade_name]
        ws = wb.create_sheet(f"2-yil {grade_name}")
        title = f"2-YIL — {grade_name.upper()} — {gdata['platforma']}"
        note = ("Haftasiga 2 dars, 42 o'quv haftasi. 84 dars — 1-yil bilan AYNAN bir xil. "
                "Dasturlash BU SINFLARDA hali yo'q (2-sinfdan boshlanadi).")
        extra = "1-yil {} dasturi bilan AYNAN bir xil (grade-anchored yondashuv). Dasturlash yo'q.".format(grade_name)
        write_grade_sheet(ws, title, note, gdata, grade_key=grade_name, extra_note=extra)

    # ---------------- 2-yil 2-3-sinf: qurish (1-yil bilan bir xil) + YANGI dasturlash (3-4-chorak) ----------------
    for grade_name in ["2-sinf", "3-sinf"]:
        gdata = ALL_GRADES_Y1[grade_name]
        ws = wb.create_sheet(f"2-yil {grade_name}")
        title = f"2-YIL — {grade_name.upper()} — {gdata['platforma']} + Dasturlash (3-4-chorak)"
        note = ("1-2-chorak: haftasiga 2 dars, 100% QURISH (1-yil bilan bir xil). 3-4-chorak: "
                "haftasiga 3 dars — 2 dars QURISH (o'zgarishsiz) + YANGI 1 dars DASTURLASH (Scratch-uslub). "
                "Dasturlash aynan shu sinfdan (2-sinf) boshlanadi.")
        extra = ("Qurish qismi 1-yil {} dasturi bilan AYNAN bir xil (grade-anchored yondashuv). "
                 "Dasturlash — 2-sinfdan boshlab qo'shiladigan YANGI kurs.").format(grade_name)
        last_row = write_grade_sheet(ws, title, note, gdata, grade_key=grade_name, extra_note=extra)
        final_row, dast_total_row = write_dasturlash_section(ws, last_row, grade_name)

    # ---------------- 2-yil 4-sinf (SPIKE, new) ----------------
    ws = wb.create_sheet("2-yil 4-sinf")
    title = f"2-YIL — 4-SINF — {GRADE4_Y2['platforma']} (YANGI PLATFORMA)"
    note = ("Haftasiga 2 dars, 42 o'quv haftasi. 84 dars: 4 chorak x 21 dars "
            "(19 mavzu darsi + 1 nazorat ishi + 1 loyiha). SPIKE Prime rasmiy darslik/missiya "
            "to'plamidan foydalaniladi (Makerzoid katalogi bu yerga tegishli emas). SPIKE'da qurish "
            "va dasturlash tabiiy ravishda birga o'tiladi (Prime metodikasi).")
    extra = "0-4-sinf Makerzoid bosqichini tugatgan o'quvchilar uchun yangi platforma — SPIKE Prime."
    write_grade_sheet(ws, title, note, GRADE4_Y2, grade_key=None, extra_note=extra)

    # ---------------- Instruksiya katalogi (appendix) ----------------
    cat = wb.create_sheet("Instruksiya katalogi")
    cat["A1"] = "MAKERZOID ROBOT MASTER (PREMIUM) — INSTRUKSIYA KATALOGI"
    cat.merge_cells("A1:B1")
    style_title(cat["A1"])
    cat.row_dimensions[1].height = 22
    cat["A2"] = ("Google Drive \"Robot master(PM) instruction\" papkasidan yig'ilgan to'liq ro'yxat "
                 f"({sum(len(v) for v in CATALOG.values())} model, {len(CATALOG)} bo'lim). "
                 "Har bir model uchun qadam-baqadam rasmli instruksiya papkada mavjud.")
    cat.merge_cells("A2:B2")
    cat["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    style_note(cat["A2"])
    cat.row_dimensions[2].height = 28

    row = 4
    cat.cell(row=row, column=1, value="Bo'lim")
    cat.cell(row=row, column=2, value="Model nomi")
    style_header(cat.cell(row=row, column=1))
    style_header(cat.cell(row=row, column=2))
    row += 1
    for section, models in CATALOG.items():
        for m in models:
            cat.cell(row=row, column=1, value=section)
            cat.cell(row=row, column=2, value=m)
            style_cell(cat.cell(row=row, column=1))
            style_cell(cat.cell(row=row, column=2))
            row += 1
    autosize(cat, [30, 40])
    cat.freeze_panes = "A5"

    # order sheets nicely
    order = ["Umumiy reja",
             "1-yil 0-sinf", "1-yil 1-sinf", "1-yil 2-sinf", "1-yil 3-sinf", "1-yil 4-sinf",
             "2-yil 0-sinf", "2-yil 1-sinf", "2-yil 2-sinf", "2-yil 3-sinf", "2-yil 4-sinf",
             "Instruksiya katalogi"]
    wb._sheets = [wb[name] for name in order]

    # output/ repo'dan tashqarida (project_bundle/output), curriculum/ esa site/ ichida:
    # site/curriculum -> site -> project_bundle -> output
    here = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.join(here, "..", "..", "output")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Dastur_0-4sinf_Makerzoid_SPIKE.xlsx")
    wb.save(out_path)
    print("Saved:", out_path)

if __name__ == "__main__":
    build()
